from flask import request, jsonify, send_file
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from models import Faculty, Student, ResearchPersonnel, Scholarship, StudentLoad, FacultyLoad, Publication
from datetime import datetime
import os
import io

def register_report_routes(app, db):
    
    @app.route('/api/reports/faculty', methods=['GET'])
    def generate_faculty_report():
        """Generate Faculty Report"""
        try:
            format_type = request.args.get('format', 'pdf').lower()
            
            # Get all faculty data
            faculty_list = Faculty.query.all()
            
            if format_type == 'pdf':
                return generate_faculty_pdf(faculty_list)
            elif format_type == 'excel':
                return generate_faculty_excel(faculty_list)
            else:
                return jsonify({'error': 'Invalid format. Use pdf or excel'}), 400
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/reports/student', methods=['GET'])
    def generate_student_report():
        """Generate Student Report"""
        try:
            format_type = request.args.get('format', 'pdf').lower()
            
            # Get all student data
            students = Student.query.all()
            
            if format_type == 'pdf':
                return generate_student_pdf(students)
            elif format_type == 'excel':
                return generate_student_excel(students)
            else:
                return jsonify({'error': 'Invalid format. Use pdf or excel'}), 400
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/reports/enrollment', methods=['GET'])
    def generate_enrollment_report():
        """Generate Student Enrollment Report"""
        try:
            format_type = request.args.get('format', 'pdf').lower()
            
            # Get enrollment data
            enrollments = db.session.query(StudentLoad).all()
            
            if format_type == 'pdf':
                return generate_enrollment_pdf(enrollments)
            elif format_type == 'excel':
                return generate_enrollment_excel(enrollments)
            else:
                return jsonify({'error': 'Invalid format. Use pdf or excel'}), 400
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/reports/scholarships', methods=['GET'])
    def generate_scholarships_report():
        """Generate Scholarships Report"""
        try:
            format_type = request.args.get('format', 'pdf').lower()
            
            scholarships = Scholarship.query.all()
            
            if format_type == 'pdf':
                return generate_scholarships_pdf(scholarships)
            elif format_type == 'excel':
                return generate_scholarships_excel(scholarships)
            else:
                return jsonify({'error': 'Invalid format. Use pdf or excel'}), 400
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/reports/research-extension', methods=['GET'])
    def generate_research_extension_report():
        """Generate Research, Extension and Publication Report"""
        try:
            format_type = request.args.get('format', 'pdf').lower()
            
            personnel = ResearchPersonnel.query.all()
            publications = Publication.query.all()
            
            if format_type == 'pdf':
                return generate_research_extension_pdf(personnel, publications)
            elif format_type == 'excel':
                return generate_research_extension_excel(personnel, publications)
            else:
                return jsonify({'error': 'Invalid format. Use pdf or excel'}), 400
        except Exception as e:
            return jsonify({'error': str(e)}), 500

def generate_faculty_pdf(faculty_list):
    """Generate Faculty Report in PDF format"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Title
    styles = getSampleStyleSheet()
    title = Paragraph("Faculty Report", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # Create table data
    table_data = [['ID', 'Name', 'Faculty ID', 'Email', 'Department', 'Position']]
    for faculty in faculty_list:
        table_data.append([
            str(faculty.id),
            faculty.name,
            faculty.faculty_id,
            faculty.email,
            faculty.department or '',
            faculty.position or ''
        ])
    
    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.3*inch))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name='Faculty_Report.pdf')

def generate_faculty_excel(faculty_list):
    """Generate Faculty Report in Excel format"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Faculty"
    
    # Headers
    headers = ['ID', 'Name', 'Faculty ID', 'Email', 'Department', 'Position', 'Phone']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Add data
    for faculty in faculty_list:
        ws.append([
            faculty.id,
            faculty.name,
            faculty.faculty_id,
            faculty.email,
            faculty.department or '',
            faculty.position or '',
            faculty.phone or ''
        ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='Faculty_Report.xlsx')

def generate_student_pdf(students):
    """Generate Student Report in PDF format"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    styles = getSampleStyleSheet()
    title = Paragraph("Student Report", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    table_data = [['ID', 'Student ID', 'Name', 'Email', 'Program', 'Year Level', 'Status']]
    for student in students:
        table_data.append([
            str(student.id),
            student.student_id,
            student.name,
            student.email,
            student.program or '',
            student.year_level or '',
            student.status
        ])
    
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.3*inch))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name='Student_Report.pdf')

def generate_student_excel(students):
    """Generate Student Report in Excel format"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    
    headers = ['ID', 'Student ID', 'Name', 'Email', 'Program', 'Year Level', 'Phone', 'Status']
    ws.append(headers)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    for student in students:
        ws.append([
            student.id,
            student.student_id,
            student.name,
            student.email,
            student.program or '',
            student.year_level or '',
            student.phone or '',
            student.status
        ])
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 18
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='Student_Report.xlsx')

def generate_enrollment_pdf(enrollments):
    """Generate Enrollment Report in PDF format"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape)
    elements = []
    
    styles = getSampleStyleSheet()
    title = Paragraph("Student Enrollment Report", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    table_data = [['Student ID', 'Academic Year', 'Semester', 'Units', 'Courses', 'GPA', 'Status']]
    for enrollment in enrollments:
        student = Student.query.get(enrollment.student_id)
        table_data.append([
            student.student_id if student else 'N/A',
            enrollment.academic_year,
            enrollment.semester,
            str(enrollment.total_units or ''),
            str(enrollment.courses_enrolled or ''),
            str(enrollment.gpa or ''),
            enrollment.status or ''
        ])
    
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name='Enrollment_Report.pdf')

def generate_enrollment_excel(enrollments):
    """Generate Enrollment Report in Excel format"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Enrollment"
    
    headers = ['Student ID', 'Name', 'Academic Year', 'Semester', 'Units', 'Courses', 'GPA', 'Status']
    ws.append(headers)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    for enrollment in enrollments:
        student = Student.query.get(enrollment.student_id)
        ws.append([
            student.student_id if student else 'N/A',
            student.name if student else 'N/A',
            enrollment.academic_year,
            enrollment.semester,
            enrollment.total_units or '',
            enrollment.courses_enrolled or '',
            enrollment.gpa or '',
            enrollment.status or ''
        ])
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 18
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='Enrollment_Report.xlsx')

def generate_scholarships_pdf(scholarships):
    """Generate Scholarships Report in PDF format"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    styles = getSampleStyleSheet()
    title = Paragraph("Scholarships Report", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    table_data = [['Student ID', 'Scholarship Name', 'Type', 'Amount', 'Academic Year', 'Semester']]
    for scholarship in scholarships:
        student = Student.query.get(scholarship.student_id)
        table_data.append([
            student.student_id if student else 'N/A',
            scholarship.scholarship_name,
            scholarship.scholarship_type or '',
            f"${scholarship.amount:.2f}" if scholarship.amount else '',
            scholarship.academic_year or '',
            scholarship.semester or ''
        ])
    
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name='Scholarships_Report.pdf')

def generate_scholarships_excel(scholarships):
    """Generate Scholarships Report in Excel format"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Scholarships"
    
    headers = ['Student ID', 'Student Name', 'Scholarship Name', 'Type', 'Amount', 'Academic Year', 'Semester']
    ws.append(headers)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    for scholarship in scholarships:
        student = Student.query.get(scholarship.student_id)
        ws.append([
            student.student_id if student else 'N/A',
            student.name if student else 'N/A',
            scholarship.scholarship_name,
            scholarship.scholarship_type or '',
            scholarship.amount or '',
            scholarship.academic_year or '',
            scholarship.semester or ''
        ])
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 20
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='Scholarships_Report.xlsx')

def generate_research_extension_pdf(personnel, publications):
    """Generate Research, Extension and Publication Report in PDF format"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    styles = getSampleStyleSheet()
    title = Paragraph("Research, Extension and Publication Report", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # Personnel section
    elements.append(Paragraph("Research Personnel", styles['Heading2']))
    table_data = [['ID', 'Name', 'Email', 'Position', 'Specialization']]
    for person in personnel:
        table_data.append([
            str(person.id),
            person.name,
            person.email,
            person.position or '',
            person.specialization or ''
        ])
    
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Publications section
    elements.append(PageBreak())
    elements.append(Paragraph("Publications", styles['Heading2']))
    pub_data = [['Title', 'Type', 'Journal', 'Publication Date']]
    for pub in publications:
        pub_data.append([
            pub.title[:40],
            pub.publication_type or '',
            pub.journal_name or '',
            pub.publication_date.strftime('%Y-%m-%d') if pub.publication_date else ''
        ])
    
    pub_table = Table(pub_data)
    pub_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    elements.append(pub_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name='Research_Extension_Publication_Report.pdf')

def generate_research_extension_excel(personnel, publications):
    """Generate Research, Extension and Publication Report in Excel format"""
    wb = Workbook()
    
    # Personnel sheet
    ws1 = wb.active
    ws1.title = "Personnel"
    
    headers = ['ID', 'Name', 'Email', 'Position', 'Specialization', 'Phone']
    ws1.append(headers)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    for person in personnel:
        ws1.append([
            person.id,
            person.name,
            person.email,
            person.position or '',
            person.specialization or '',
            person.phone or ''
        ])
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws1.column_dimensions[col].width = 20
    
    # Publications sheet
    ws2 = wb.create_sheet("Publications")
    headers2 = ['Title', 'Type', 'Journal', 'Volume', 'Pages', 'Publication Date', 'URL']
    ws2.append(headers2)
    
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    for pub in publications:
        ws2.append([
            pub.title,
            pub.publication_type or '',
            pub.journal_name or '',
            pub.volume or '',
            pub.pages or '',
            pub.publication_date.strftime('%Y-%m-%d') if pub.publication_date else '',
            pub.url or ''
        ])
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws2.column_dimensions[col].width = 20
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='Research_Extension_Publication_Report.xlsx')

from reportlab.lib.pagesizes import landscape
